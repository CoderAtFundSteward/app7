import csv
import io
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

from app.database.supabase import get_supabase_client

DATE_HEADERS = {"date", "transaction_date", "txn_date", "posting_date", "transactiondate"}
DESCRIPTION_HEADERS = {"description", "memo", "details", "transaction_description", "narration"}
AMOUNT_HEADERS = {"amount", "total", "transaction_amount", "value"}
BALANCE_HEADERS = {"balance", "running_balance"}
ACCOUNT_HEADERS = {"account", "account_name", "category", "ledger"}
CURRENCY_HEADERS = {"currency", "curr", "ccy"}


def _normalize_header(header: Any) -> str:
    return str(header or "").strip().lower().replace(" ", "_")


def _parse_amount(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).replace(",", "").strip()
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _parse_date(value: Any) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def _parse_csv(file_bytes: bytes) -> list[dict[str, Any]]:
    decoded = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(decoded))
    rows: list[dict[str, Any]] = []
    for row in reader:
        rows.append({(_normalize_header(k)): v for k, v in row.items()})
    return rows


def _parse_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []

    normalized_headers = [_normalize_header(h) for h in headers]
    rows: list[dict[str, Any]] = []
    for row in rows_iter:
        row_map = {}
        for idx, value in enumerate(row):
            key = normalized_headers[idx] if idx < len(normalized_headers) else f"col_{idx}"
            row_map[key] = value
        rows.append(row_map)
    return rows


def _pick(row: dict[str, Any], candidates: set[str]) -> Any:
    for key, value in row.items():
        if _normalize_header(key) in candidates:
            return value
    return None


def _normalize_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in raw_rows:
        amount = _parse_amount(_pick(row, AMOUNT_HEADERS))
        balance_amount = _parse_amount(_pick(row, BALANCE_HEADERS))
        description = _pick(row, DESCRIPTION_HEADERS)
        if amount is None and not description:
            continue

        normalized.append(
            {
                "transaction_date": _parse_date(_pick(row, DATE_HEADERS)),
                "description": str(description).strip() if description else None,
                "amount": float(amount) if amount is not None else 0.0,
                "balance": float(balance_amount) if balance_amount is not None else None,
                "currency": str(_pick(row, CURRENCY_HEADERS) or "USD").strip(),
                "account_name": str(_pick(row, ACCOUNT_HEADERS) or "").strip() or None,
                "raw_row": row,
            }
        )
    return normalized


def import_spreadsheet(member_id: str, file_name: str, file_bytes: bytes) -> dict[str, Any]:
    file_name_lower = file_name.lower()
    if file_name_lower.endswith(".csv"):
        raw_rows = _parse_csv(file_bytes)
        file_type = "csv"
    elif file_name_lower.endswith(".xlsx"):
        raw_rows = _parse_xlsx(file_bytes)
        file_type = "xlsx"
    else:
        raise ValueError("Unsupported file type. Please upload a .csv or .xlsx file.")

    normalized_rows = _normalize_rows(raw_rows)
    if not normalized_rows:
        raise ValueError("No valid transaction rows found in spreadsheet.")

    supabase = get_supabase_client()
    upload_insert = (
        supabase.table("spreadsheet_uploads")
        .insert(
            {
                "member_id": member_id,
                "file_name": file_name,
                "file_type": file_type,
                "row_count": len(normalized_rows),
                "raw_rows": raw_rows,
            }
        )
        .execute()
    )
    upload_rows = upload_insert.data or []
    if not upload_rows:
        raise RuntimeError("Failed to save spreadsheet upload.")
    upload_id = upload_rows[0]["id"]

    transactions_payload = [
        {
            "member_id": member_id,
            "upload_id": upload_id,
            **row,
        }
        for row in normalized_rows
    ]
    supabase.table("spreadsheet_transactions").insert(transactions_payload).execute()

    return {
        "upload_id": upload_id,
        "file_name": file_name,
        "file_type": file_type,
        "rows_imported": len(normalized_rows),
    }


def list_spreadsheet_uploads(member_id: str, limit: int = 20) -> list[dict[str, Any]]:
    response = (
        get_supabase_client()
        .table("spreadsheet_uploads")
        .select("id,file_name,file_type,row_count,uploaded_at")
        .eq("member_id", member_id)
        .order("uploaded_at", desc=True)
        .limit(limit)
        .execute()
    )
    return response.data or []
